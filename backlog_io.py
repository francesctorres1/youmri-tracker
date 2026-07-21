# -*- coding: utf-8 -*-
"""Lectura de los Excel de Backlog y Priorización de YouMRI."""
import pandas as pd
import openpyxl


def parse_backlog(file) -> pd.DataFrame:
    """Lee la hoja 'Backlog'. Cabecera en la fila 4 (índice 3)."""
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb["Backlog"]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 4 and row[0] and str(row[0]).startswith("US"):
            rows.append(row)
    cols = ["ID", "Épica", "Fase", "Clasificación", "Prioridad", "Rol",
            "Historia", "Criterios", "SP", "Dependencias", "Flags", "RefDoc"]
    df = pd.DataFrame(rows, columns=cols[:len(rows[0])] if rows else cols)
    # normalizar tipos a texto
    for c in df.columns:
        df[c] = df[c].apply(lambda x: "" if x is None else str(x))
    return df


def parse_priorizacion(file) -> pd.DataFrame:
    """Lee las hojas 'FASE x - Prioridad' y junta ID + 'Tu prioridad' + Notas."""
    wb = openpyxl.load_workbook(file, data_only=True)
    out = []
    for sh in wb.sheetnames:
        if "Prioridad" not in sh:
            continue
        ws = wb[sh]
        header = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if row and row[0] == "ID":
                header = list(row)
                continue
            if header and row[0] and str(row[0]).startswith("US"):
                d = dict(zip(header, row))
                out.append({
                    "ID": d.get("ID"),
                    "Prio_Francesc": d.get("Tu prioridad (1–5)"),
                    "Notas": d.get("Notas"),
                })
    return pd.DataFrame(out)
